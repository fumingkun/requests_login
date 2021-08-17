#fumingkun
#python3.8.2
from openpyxl import load_workbook

class MyExcel:

    def __init__(self,excel_path,sheet_name):
        #打开一个workbook
        wb = load_workbook(excel_path)
        #选择一个表单   通过表单名  Sheet1
        self.sh = wb[sheet_name]

    def read_data(self):
        #注意：接口的请求数据，读取出来是字符串
        #存储表单下读取的所有数据，每一个成员都是一个字典
        all_data = []
        data = list(self.sh.values)
        header = data[0]    #获取所有的列名
        for row in data[1:]:
            row_dict = dict(zip(header, row))
            all_data.append(row_dict)
        return all_data


if __name__  =="__main__":
    #excel文件路径
    excel_path = r"D:\requests_login\test_excel\cases_1.xlsx"
    my = MyExcel(excel_path,"注册、登陆接口（游）")
    cases = my.read_data()
    for case in cases:
        print(case)

    # wb = load_workbook(excel_path)
    # sh = wb["订单列表"]




