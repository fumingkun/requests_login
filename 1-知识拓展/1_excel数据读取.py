#fumingkun
#python3.8.2
"""
1、安装    pip install openpyxl
row  行
column 列




"""
from openpyxl import load_workbook

excel_path = r"D:\requests_pytest\test_excel\cases.xlsx"
# 加载一个excel，获取一个工作簿
wb = load_workbook(excel_path)

# 选择一个表单 - 通过表单名
sh = wb["订单列表"]

# 在选择的表单当中，读取某个单元格的数据，修改/写入数据到某个单元格  cell
# 行号和列号都是从1开始的
# cell_value = sh.cell(2,3).value
# print(cell_value)

# 搭档当前sheet的总行号，总列号
row_nams = sh.max_row
col_nams = sh.max_column

# 读取第一行（作为key）,行号是1，
# keys = []
# for col_index in range(1,sh.max_column + 1):
#     keys.append(sh.cell(1,col_index).value)
# # print(keys)
#
# # 获取行号，取第一行
# for row_index in range(2,sh.max_row + 1):
#     values = []
#     # 在每一行里面，从第一列开始，获取所有列的值
#     for col_index in range(1, sh.max_column + 1):
#         values.append(sh.cell(row_index,col_index).value)
#     case = dict(zip(keys,values))
#     # keys和values打包， -zip函数
#     print(case)

# 第二种方法
data = list(sh.values)
header = data[0]
all_data = []
for row in data[1:]:
    row_dict = dict(zip(header,row))
    all_data.append(row_dict)
print(all_data)


# 读取所有行
# for row in sh.rows:
#     # print(row)
#     for item in row:
#         print(item.value,end=" ")
#     print()

# 给某个单元格写入值
# sh.cell(2,3).value = "post"
# # 做了修改，就要保存
# # filename如果不是打开的excel文件，那就是另存为，如果是打开的excel文件，保存到原文件中
# # 保存时，要保证没有其他程序在使用当前文件，否是会报错，PermissionError
# wb.save(excel_path)









